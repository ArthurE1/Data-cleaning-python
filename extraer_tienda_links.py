from pathlib import Path
import re
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

# ====== CONFIGURA AQUÍ ======
RUTA = Path(r"D:\PythonExcel\20250825_entregado_detergentes.xlsx")  # <-- tu archivo
NOMBRE_HOJA = None   # None = primera hoja; o pon el nombre exacto
COL_TIENDA = "E"     # columna TIENDA
COL_LINK = "L"       # columna id visita / link
FILA_INICIO = 2      # asumo fila 1 = encabezados

# Si en L hay solo GUID (36 chars) y quieres armar la URL:
URL_PREFIX = "https://services.traxretail.com/trax-one/sindicatedmx/explore/visit/"
# ============================

GUID_RE = re.compile(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$")

def extraer_url_de_formula(s: str) -> str | None:
    # Soporta =HIPERVINCULO("url","texto") o =HYPERLINK("url","text")
    # Captura la PRIMERA URL entre comillas
    m = re.search(r'"(https?://[^"]+)"', s, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    return None

def main():
    # IMPORTANTE: data_only=False para ver la fórmula HIPERVINCULO
    wb = load_workbook(RUTA, data_only=False, keep_links=True)
    ws = wb[NOMBRE_HOJA] if NOMBRE_HOJA else wb.worksheets[0]

    idx_tienda = column_index_from_string(COL_TIENDA)
    idx_link = column_index_from_string(COL_LINK)

    pares = []  # (tienda, url)
    diag = []   # diagnóstico para revisar qué leyó

    for row in ws.iter_rows(min_row=FILA_INICIO, values_only=False):
        tienda = row[idx_tienda - 1].value
        celda = row[idx_link - 1]

        url = None
        # a) Si hay hipervínculo "real" (Insertar vínculo)
        if celda.hyperlink and celda.hyperlink.target:
            url = celda.hyperlink.target

        # b) Si hay fórmula HIPERVINCULO/HYPERLINK, léela (con data_only=False)
        if not url and isinstance(celda.value, str) and celda.value.startswith(("=HIPERVINCULO(", "=HYPERLINK(")):
            url = extraer_url_de_formula(celda.value)

        # c) Si solo hay GUID, arma la URL
        if not url and isinstance(celda.value, str) and GUID_RE.fullmatch(celda.value.strip()):
            url = URL_PREFIX + celda.value.strip()

        if tienda and url:
            pares.append((str(tienda).strip(), url.strip()))

        # Para revisar qué hay en esa celda
        diag.append({
            "tienda": tienda,
            "valor_celda_L": celda.value,
            "hyperlink_target": (celda.hyperlink.target if celda.hyperlink else None),
            "url_detectada": url
        })

    df = pd.DataFrame(pares, columns=["tienda", "link"])
    df_unique = df.drop_duplicates(subset=["tienda", "link"]).sort_values(["tienda", "link"])

    # Agrupar links por tienda
    agrupado = defaultdict(list)
    for t, l in df_unique.itertuples(index=False):
        agrupado[t].append(l)
    df_group = pd.DataFrame([(t, "\n".join(links)) for t, links in sorted(agrupado.items())],
                            columns=["tienda", "links_de_tienda"])

    # Diagnóstico: primeras 50 filas
    df_diag = pd.DataFrame(diag).head(50)

    salida = RUTA.with_name(RUTA.stem + "_tienda_links.xlsx")
    with pd.ExcelWriter(salida, engine="openpyxl") as xw:
        df_unique.to_excel(xw, index=False, sheet_name="pares_unicos")
        df_group.to_excel(xw, index=False, sheet_name="links_por_tienda")
        df_diag.to_excel(xw, index=False, sheet_name="diagnostico")

    print(f"Filas con URL extraída: {len(df)} | Únicas (tienda,link): {len(df_unique)}")
    print(f"Listo -> {salida}")

if __name__ == "__main__":
    main()
